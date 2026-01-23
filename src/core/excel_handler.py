"""
Excel Handler for Academic Evaluation System
Manages multi-sheet Excel workbooks with academic data.

COMPONENT STATUS: ✅ COMPLETE
LAST UPDATED: 2025-01-21
DEPENDENCIES: pandas, openpyxl, config.settings, utils.logger
"""
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from loguru import logger
import json
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from config.settings import EXCEL_DIR, EXCEL_FILENAME, EXCEL_SHEET_NAME, EXCEL_COURSES_SHEET_NAME


class ExcelHandler:
    """Handles Excel file creation, data appending, and formatting for academic data.
    
    Features:
    - Multi-sheet workbooks (Student Data + Course Details)
    - Batch file management with timestamps
    - Auto-formatting (headers, filters, column widths)
    - Metadata tracking
    """
    
    def __init__(self):
        """Initialize Excel handler with default file paths."""
        self.excel_dir = EXCEL_DIR
        self.excel_file = self.excel_dir / EXCEL_FILENAME
        self.sheet_name = EXCEL_SHEET_NAME
        self.courses_sheet_name = EXCEL_COURSES_SHEET_NAME
        self.batch_metadata_file = self.excel_dir / "batch_metadata.json"
        
        # Ensure directory exists
        self.excel_dir.mkdir(exist_ok=True)
        
        # Initialize batch metadata
        self._load_batch_metadata()
        
        logger.info(f"Excel Handler initialized with file: {self.excel_file}")
    
    def _load_batch_metadata(self):
        """Load batch metadata from JSON file."""
        if self.batch_metadata_file.exists():
            try:
                with open(self.batch_metadata_file, 'r') as f:
                    self.batch_metadata = json.load(f)
            except Exception as e:
                logger.error(f"Failed to load batch metadata: {e}")
                self.batch_metadata = {"batches": [], "current_batch": None}
        else:
            self.batch_metadata = {"batches": [], "current_batch": None}
    
    def _save_batch_metadata(self):
        """Save batch metadata to JSON file."""
        try:
            with open(self.batch_metadata_file, 'w') as f:
                json.dump(self.batch_metadata, f, indent=2)
        except Exception as e:
            logger.error(f"Failed to save batch metadata: {e}")
    
    def create_batch_excel_file(self, batch_name: str = None) -> tuple[bool, str]:
        """Create a new Excel file with timestamp for a batch.
        
        Args:
            batch_name: Optional custom batch name
            
        Returns:
            Tuple of (success, batch_filename)
        """
        try:
            # Generate batch name with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if batch_name:
                batch_filename = f"academic_batch_{batch_name}_{timestamp}.xlsx"
            else:
                batch_filename = f"academic_batch_{timestamp}.xlsx"
            
            batch_file_path = self.excel_dir / batch_filename
            
            # Define column headers for Student Data sheet
            headers = [
                'Timestamp',
                'Document Filename',
                'Student Name',
                'Roll Number',
                'Email',
                'Phone',
                'Department',
                'Program',
                'Semester',
                'Academic Year',
                'CGPA',
                'SGPA',
                'Attendance Percentage',
                'Date of Birth',
                'Gender',
                'Category',
                'Awards and Honors',
                'Extracurricular Activities',
                'Remarks',
                'Model Used',
                'Tokens Used',
                'Analysis Status'
            ]
            
            # Create DataFrame with headers
            df = pd.DataFrame(columns=headers)
            
            # Save to Excel with formatting
            with pd.ExcelWriter(batch_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=self.sheet_name, index=False)
                
                # Get the worksheet
                worksheet = writer.sheets[self.sheet_name]
                
                # Apply formatting
                self._format_excel_sheet(worksheet, headers)
            
            # Create courses sheet (Sheet2)
            self.create_courses_sheet(batch_file_path)
            
            # Update batch metadata
            batch_info = {
                "filename": batch_filename,
                "created_at": datetime.now().isoformat(),
                "file_path": str(batch_file_path),
                "record_count": 0
            }
            
            self.batch_metadata["batches"].append(batch_info)
            self.batch_metadata["current_batch"] = batch_filename
            
            # Keep only last 10 batches
            if len(self.batch_metadata["batches"]) > 10:
                oldest_batch = self.batch_metadata["batches"].pop(0)
                try:
                    oldest_file = Path(oldest_batch["file_path"])
                    if oldest_file.exists():
                        oldest_file.unlink()
                        logger.info(f"Removed old batch file: {oldest_file}")
                except Exception as e:
                    logger.error(f"Failed to remove old batch file: {e}")
            
            self._save_batch_metadata()
            
            logger.info(f"Created new batch Excel file: {batch_file_path}")
            return True, batch_filename
            
        except Exception as e:
            logger.error(f"Failed to create batch Excel file: {e}")
            return False, ""
    
    def create_courses_sheet(self, batch_file_path: Path):
        """Create Course Details sheet (Sheet2) for course-wise data.
        
        Args:
            batch_file_path: Path to the batch Excel file
        """
        try:
            # Define courses sheet headers
            courses_headers = [
                'Timestamp',
                'Document Filename',
                'Student Name',
                'Roll Number',
                'Course Code',
                'Course Name',
                'Credits',
                'Grade',
                'Semester',
                'Academic Year'
            ]
            
            # Load existing workbook
            try:
                workbook = openpyxl.load_workbook(batch_file_path)
            except Exception:
                workbook = openpyxl.Workbook()
            
            # Create Sheet2
            if self.courses_sheet_name in workbook.sheetnames:
                workbook.remove(workbook[self.courses_sheet_name])
            
            courses_sheet = workbook.create_sheet(title=self.courses_sheet_name, index=1)
            
            # Write headers
            for col, header in enumerate(courses_headers, 1):
                courses_sheet.cell(row=1, column=col, value=header)
            
            # Format courses sheet
            self._format_excel_sheet(courses_sheet, courses_headers)
            
            # Save workbook
            workbook.save(batch_file_path)
            logger.info(f"Created courses sheet '{self.courses_sheet_name}' in {batch_file_path}")
            
        except Exception as e:
            logger.error(f"Failed to create courses sheet: {e}")
    
    def _format_excel_sheet(self, worksheet, headers: List[str]):
        """Apply formatting to the Excel sheet.
        
        Args:
            worksheet: The worksheet to format
            headers: List of column headers
        """
        try:
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply header formatting
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for col_num, header in enumerate(headers, 1):
                column_letter = openpyxl.utils.get_column_letter(col_num)
                max_length = len(header)
                
                # Set minimum width
                worksheet.column_dimensions[column_letter].width = max(max_length + 2, 15)
            
            # Freeze header row
            worksheet.freeze_panes = "A2"
            
            logger.info("Applied formatting to Excel sheet")
            
        except Exception as e:
            logger.error(f"Failed to format Excel sheet: {e}")
    
    def append_data_to_batch(
        self, 
        analysis_data: Dict[str, Any], 
        document_filename: str, 
        batch_filename: str = None
    ) -> bool:
        """Append analysis data to a specific batch Excel file.
        
        Args:
            analysis_data: Dictionary containing analysis results
            document_filename: Name of the PDF/document analyzed
            batch_filename: Name of the batch file (uses current batch if None)
            
        Returns:
            True if data was appended successfully, False otherwise
        """
        try:
            # Use current batch if not specified
            if not batch_filename:
                batch_filename = self.batch_metadata.get("current_batch")
                if not batch_filename:
                    logger.error("No current batch file available")
                    return False
            
            batch_file_path = self.excel_dir / batch_filename
            
            if not batch_file_path.exists():
                logger.error(f"Batch file does not exist: {batch_file_path}")
                return False
            
            # Prepare data row
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Extract metadata
            metadata = analysis_data.get('_metadata', {})
            
            # Create data row
            data_row = {
                'Timestamp': current_time,
                'Document Filename': document_filename,
                'Student Name': analysis_data.get('Student Name', 'N/A'),
                'Roll Number': analysis_data.get('Roll Number', 'N/A'),
                'Email': analysis_data.get('Email', 'N/A'),
                'Phone': analysis_data.get('Phone', 'N/A'),
                'Department': analysis_data.get('Department', 'N/A'),
                'Program': analysis_data.get('Program', 'N/A'),
                'Semester': analysis_data.get('Semester', 'N/A'),
                'Academic Year': analysis_data.get('Academic Year', 'N/A'),
                'CGPA': analysis_data.get('CGPA', 'N/A'),
                'SGPA': analysis_data.get('SGPA', 'N/A'),
                'Attendance Percentage': analysis_data.get('Attendance Percentage', 'N/A'),
                'Date of Birth': analysis_data.get('Date of Birth', 'N/A'),
                'Gender': analysis_data.get('Gender', 'N/A'),
                'Category': analysis_data.get('Category', 'N/A'),
                'Awards and Honors': analysis_data.get('Awards and Honors', 'N/A'),
                'Extracurricular Activities': analysis_data.get('Extracurricular Activities', 'N/A'),
                'Remarks': analysis_data.get('Remarks', 'N/A'),
                'Model Used': metadata.get('model', 'N/A'),
                'Tokens Used': metadata.get('total_tokens', 'N/A'),
                'Analysis Status': 'Success' if not metadata.get('error') else 'Error'
            }
            
            # Read existing data
            try:
                existing_df = pd.read_excel(batch_file_path, sheet_name=self.sheet_name)
            except Exception:
                existing_df = pd.DataFrame()
            
            # Check for duplicates
            if not existing_df.empty and 'Document Filename' in existing_df.columns:
                if document_filename in existing_df['Document Filename'].values:
                    logger.warning(f"Document {document_filename} already exists in batch, skipping")
                    return True
            
            # Append new data
            new_df = pd.DataFrame([data_row])
            updated_df = pd.concat([existing_df, new_df], ignore_index=True)
            
            # Save updated data
            with pd.ExcelWriter(batch_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                updated_df.to_excel(writer, sheet_name=self.sheet_name, index=False)
            
            # Update batch metadata
            self._update_batch_record_count(batch_filename, len(updated_df))
            
            logger.info(f"Appended data for {document_filename} to batch file {batch_filename}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to append data to batch file: {e}")
            return False
    
    def append_courses_data(
        self,
        courses_data: List[Dict],
        student_name: str,
        roll_number: str,
        document_filename: str,
        batch_filename: str = None
    ) -> bool:
        """Append course data to Course Details sheet.
        
        Args:
            courses_data: List of course dictionaries
            student_name: Student's name
            roll_number: Student's roll number
            document_filename: Document filename
            batch_filename: Batch filename (uses current if None)
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not courses_data or not isinstance(courses_data, list):
                logger.info(f"No courses data to append for {student_name}")
                return True
            
            # Use current batch if not specified
            if not batch_filename:
                batch_filename = self.batch_metadata.get("current_batch")
            
            batch_file_path = self.excel_dir / batch_filename
            
            # Read existing courses data
            try:
                df = pd.read_excel(batch_file_path, sheet_name=self.courses_sheet_name)
            except Exception:
                df = pd.DataFrame()
            
            # Prepare course rows
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            new_rows = []
            
            for course in courses_data:
                course_row = {
                    'Timestamp': current_time,
                    'Document Filename': document_filename,
                    'Student Name': student_name,
                    'Roll Number': roll_number,
                    'Course Code': course.get('Course Code', 'N/A'),
                    'Course Name': course.get('Course Name', 'N/A'),
                    'Credits': course.get('Credits', 'N/A'),
                    'Grade': course.get('Grade', 'N/A'),
                    'Semester': course.get('Semester', 'N/A'),
                    'Academic Year': course.get('Academic Year', 'N/A')
                }
                new_rows.append(course_row)
            
            # Append new courses
            new_df = pd.DataFrame(new_rows)
            df = pd.concat([df, new_df], ignore_index=True)
            
            # Save to Excel
            with pd.ExcelWriter(batch_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=self.courses_sheet_name, index=False)
            
            logger.info(f"Appended {len(courses_data)} courses for {student_name}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to append courses data: {e}")
            return False
    
    def _update_batch_record_count(self, batch_filename: str, record_count: int):
        """Update record count for a batch in metadata."""
        for batch in self.batch_metadata["batches"]:
            if batch["filename"] == batch_filename:
                batch["record_count"] = record_count
                break
        self._save_batch_metadata()
    
    def get_current_batch(self) -> Optional[str]:
        """Get current batch filename."""
        return self.batch_metadata.get("current_batch")
    
    def get_available_batches(self) -> List[Dict[str, Any]]:
        """Get list of available batch files."""
        return self.batch_metadata.get("batches", [])
    
    def get_batch_file_path(self, batch_filename: str) -> Path:
        """Get full path for a batch filename."""
        return self.excel_dir / batch_filename

class EnhancedExcelHandler(ExcelHandler):
    """Excel handler with TASK_SPECIFICATIONS.md structure"""
    
    def create_enhanced_workbook(self, students: List[dict], 
                                 batch_name: str) -> str:
        """
        Create 3-sheet workbook as per specifications.
        
        Sheets:
        1. Raw_Records - All extracted data
        2. Subject_Performance - Course statistics
        3. At_Risk_Students - Intervention list
        """
        wb = openpyxl.Workbook()
        
        # Sheet 1: Raw Records
        ws1 = wb.active
        ws1.title = "Raw_Records"
        self._create_raw_records_sheet(ws1, students)
        
        # Sheet 2: Subject Performance
        ws2 = wb.create_sheet("Subject_Performance")
        self._create_subject_sheet(ws2, students)
        
        # Sheet 3: At-Risk Students
        ws3 = wb.create_sheet("At_Risk_Students")
        self._create_at_risk_sheet(ws3, students)
        
        # Save
        filename = f"academic_evaluation_{batch_name}.xlsx"
        filepath = Path(EXCEL_DIR) / filename
        wb.save(filepath)
        
        return str(filepath)
    
    def _create_raw_records_sheet(self, ws, students):
        """Sheet 1: Raw extracted records"""
        # Headers
        headers = [
            "Record_ID", "Student_Name", "Roll_Number", "Email",
            "Department", "Program", "Semester", "CGPA", "SGPA",
            "Attendance_Percentage", "Credits_Earned", 
            "Credits_Registered", "Backlogs_Count", 
            "Current_Standing", "Confidence_Score", "Requires_Review"
        ]
        
        # Write headers with formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", 
                                   end_color="4472C4", 
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for idx, student in enumerate(students, 2):
            ws.cell(idx, 1, idx-1)  # Record_ID
            ws.cell(idx, 2, student.get('student_name'))
            ws.cell(idx, 3, student.get('roll_number'))
            ws.cell(idx, 4, student.get('email'))
            ws.cell(idx, 5, student.get('department'))
            ws.cell(idx, 6, student.get('program', 'B.Tech'))
            ws.cell(idx, 7, student.get('semester'))
            ws.cell(idx, 8, student.get('cgpa'))
            ws.cell(idx, 9, student.get('sgpa'))
            ws.cell(idx, 10, student.get('attendance_percentage'))
            ws.cell(idx, 11, student.get('credits_earned'))
            ws.cell(idx, 12, student.get('credits_registered'))
            ws.cell(idx, 13, student.get('backlogs_count', 0))
            ws.cell(idx, 14, student.get('current_standing'))
            
            # Confidence score
            confidence = student.get('confidence_metadata', {}).get('overall_confidence', 1.0)
            ws.cell(idx, 15, round(confidence, 2))
            
            # Highlight low confidence
            if confidence < 0.8:
                for col in range(1, 17):
                    ws.cell(idx, col).fill = PatternFill(
                        start_color="FFC7CE",
                        end_color="FFC7CE",
                        fill_type="solid"
                    )
            
            ws.cell(idx, 16, "TRUE" if confidence < 0.8 else "FALSE")
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    def _create_subject_sheet(self, ws, students):
        """Sheet 2: Subject-wise performance"""
        from src.core.dashboard_analytics import DashboardAnalytics
        
        subjects = DashboardAnalytics.calculate_subject_averages(students)
        
        # Headers
        headers = [
            "Course_Code", "Course_Name", "Students_Enrolled",
            "Avg_Grade_Points", "Pass_Percentage", "Difficulty_Level"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47",
                                   end_color="70AD47",
                                   fill_type="solid")
        
        # Write data
        for idx, (code, data) in enumerate(subjects.items(), 2):
            ws.cell(idx, 1, code)
            ws.cell(idx, 2, data['name'])
            ws.cell(idx, 3, data['count'])
            ws.cell(idx, 4, data['average'])
            ws.cell(idx, 5, 100.0)  # Placeholder
            ws.cell(idx, 6, data['difficulty'])
            
            # Color code difficulty
            if data['difficulty'] == "Difficult":
                ws.cell(idx, 6).font = Font(color="C00000", bold=True)
        
        ws.freeze_panes = "A2"
    
    def _create_at_risk_sheet(self, ws, students):
        """Sheet 3: At-risk students"""
        from src.core.dashboard_analytics import DashboardAnalytics
        
        at_risk = DashboardAnalytics.identify_at_risk_students(students)
        
        # Headers
        headers = [
            "Priority", "Student_Name", "Roll_Number", "Department",
            "CGPA", "Attendance", "Risk_Factors", "Action_Required"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000",
                                   end_color="C00000",
                                   fill_type="solid")
        
        # Write data
        for idx, student in enumerate(at_risk, 2):
            ws.cell(idx, 1, student['priority'])
            ws.cell(idx, 2, student['name'])
            ws.cell(idx, 3, student['roll'])
            ws.cell(idx, 4, student['dept'])
            ws.cell(idx, 5, student['cgpa'])
            ws.cell(idx, 6, student['attendance'])
            ws.cell(idx, 7, ", ".join(student['risks']))
            ws.cell(idx, 8, "Immediate intervention" if student['priority'] == "High" else "Monitor")
            
            # Highlight high priority
            if student['priority'] == "High":
                for col in range(1, 9):
                    ws.cell(idx, col).fill = PatternFill(
                        start_color="FFC7CE",
                        end_color="FFC7CE",
                        fill_type="solid"
                    )
        
        ws.freeze_panes = "A2"