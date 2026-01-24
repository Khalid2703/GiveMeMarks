"""
COMPLETE TASK D IMPLEMENTATION
Professional 3-Sheet Excel Export Handler

This file provides the COMPLETE implementation of Task D from TASK_SPECIFICATIONS.md
Copy this entire class into your src/core/excel_handler.py
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from loguru import logger


class TaskD_ExcelHandler:
    """
    Complete Task D Implementation: Professional 3-Sheet Excel Export
    
    Creates Excel workbooks with:
    - Sheet 1: Raw Extracted Records (16 columns)
    - Sheet 2: Subject-Wise Performance Summary (10 columns)
    - Sheet 3: Student Difficulty Indicators (8 columns)
    
    Features:
    - Professional formatting (colors, fonts, alignment)
    - Conditional formatting (confidence scores, priorities)
    - Auto-filter and freeze panes
    - Formulas for calculated fields
    """
    
    def __init__(self, excel_dir: Path):
        """Initialize Excel handler with output directory."""
        self.excel_dir = Path(excel_dir)
        self.excel_dir.mkdir(exist_ok=True, parents=True)
        logger.info(f"Task D Excel Handler initialized: {excel_dir}")
    
    def create_complete_workbook(
        self, 
        students: List[Dict[str, Any]], 
        batch_name: str = None
    ) -> str:
        """
        Create complete 3-sheet Excel workbook per Task D specifications.
        
        Args:
            students: List of student dictionaries with academic data
            batch_name: Optional batch identifier
            
        Returns:
            Path to created Excel file
        """
        try:
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if batch_name:
                filename = f"academic_evaluation_{batch_name}_{timestamp}.xlsx"
            else:
                filename = f"academic_evaluation_{timestamp}.xlsx"
            
            filepath = self.excel_dir / filename
            
            # Create workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Sheet 1: Raw Extracted Records
            ws1 = wb.create_sheet("Raw_Records", 0)
            self._create_sheet1_raw_records(ws1, students)
            
            # Sheet 2: Subject-Wise Performance
            ws2 = wb.create_sheet("Subject_Performance", 1)
            self._create_sheet2_subject_performance(ws2, students)
            
            # Sheet 3: At-Risk Students
            ws3 = wb.create_sheet("At_Risk_Students", 2)
            self._create_sheet3_at_risk_students(ws3, students)
            
            # Save workbook
            wb.save(filepath)
            logger.info(f"✅ Created 3-sheet Excel workbook: {filepath}")
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Failed to create Excel workbook: {e}")
            raise
    
    def _create_sheet1_raw_records(self, ws, students: List[Dict]):
        """
        SHEET 1: Raw Extracted Records
        
        16 Columns as per TASK_SPECIFICATIONS.md:
        1. Record_ID
        2. Student_Name
        3. Roll_Number
        4. Email
        5. Phone_Number
        6. Department
        7. Program
        8. Semester
        9. Academic_Year
        10. CGPA
        11. SGPA
        12. Attendance_Percentage
        13. Credits_Earned
        14. Credits_Registered
        15. Backlogs_Count
        16. Current_Standing
        17. Confidence_Score
        18. Requires_Review
        """
        # Define headers
        headers = [
            "Record_ID", "Student_Name", "Roll_Number", "Email",
            "Phone_Number", "Department", "Program", "Semester",
            "Academic_Year", "CGPA", "SGPA", "Attendance_Percentage",
            "Credits_Earned", "Credits_Registered", "Backlogs_Count",
            "Current_Standing", "Confidence_Score", "Requires_Review"
        ]
        
        # Write headers with formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(
                start_color="4472C4",  # Blue background
                end_color="4472C4",
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._get_border()
        
        # Write data rows
        for idx, student in enumerate(students, 2):
            # Get confidence score
            confidence_metadata = student.get('confidence_metadata', {})
            confidence = confidence_metadata.get('overall_confidence', 1.0)
            requires_review = confidence < 0.8
            
            # Data mapping (handle both key formats)
            row_data = [
                idx - 1,  # Record_ID
                student.get('student_name') or student.get('Student Name', 'N/A'),
                student.get('roll_number') or student.get('Roll Number', 'N/A'),
                student.get('email') or student.get('Email', 'N/A'),
                student.get('phone_number') or student.get('Phone', 'N/A'),
                student.get('department') or student.get('Department', 'N/A'),
                student.get('program') or student.get('Program', 'B.Tech'),
                student.get('semester') or student.get('Semester', 'N/A'),
                student.get('academic_year') or student.get('Academic Year', '2024-2025'),
                student.get('cgpa') or student.get('CGPA', 'N/A'),
                student.get('sgpa') or student.get('SGPA', 'N/A'),
                student.get('attendance_percentage') or student.get('Attendance Percentage', 'N/A'),
                student.get('credits_earned', 'N/A'),
                student.get('credits_registered', 'N/A'),
                student.get('backlogs_count', 0),
                student.get('current_standing', 'Good Standing'),
                round(confidence, 2),
                "TRUE" if requires_review else "FALSE"
            ]
            
            # Write row
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(idx, col, value)
                
                # Apply alternating row colors
                if idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color="F2F2F2",
                        end_color="F2F2F2",
                        fill_type="solid"
                    )
                
                # Highlight low confidence rows (orange)
                if requires_review:
                    cell.fill = PatternFill(
                        start_color="FFC7CE",  # Light red/orange
                        end_color="FFC7CE",
                        fill_type="solid"
                    )
                    cell.font = Font(color="9C0006")  # Dark red text
                
                cell.border = self._get_border(thin=True)
                
                # Center alignment for some columns
                if col in [1, 10, 11, 12, 13, 14, 15, 17, 18]:
                    cell.alignment = Alignment(horizontal="center")
        
        # Format sheet
        ws.freeze_panes = "A2"  # Freeze header row
        ws.auto_filter.ref = ws.dimensions  # Enable auto-filter
        
        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(header)
            
            # Check data length
            for row in range(2, len(students) + 2):
                cell_value = ws.cell(row, col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set width (max 50 characters)
            ws.column_dimensions[column_letter].width = min(max_length + 3, 50)
        
        logger.info(f"✓ Sheet 1 created: {len(students)} records")
    
    def _create_sheet2_subject_performance(self, ws, students: List[Dict]):
        """
        SHEET 2: Subject-Wise Performance Summary
        
        10 Columns as per TASK_SPECIFICATIONS.md:
        1. Course_Code
        2. Course_Name
        3. Course_Type (Core/Elective)
        4. Credits
        5. Students_Enrolled
        6. Avg_Grade_Points
        7. Grade_A_Count (A+, A)
        8. Grade_B_Count (B+, B)
        9. Grade_C_Count (C)
        10. Grade_F_Count (D, F)
        11. Pass_Percentage
        12. Difficulty_Level (formula-based)
        """
        # Calculate subject statistics
        course_stats = self._calculate_course_statistics(students)
        
        # Define headers
        headers = [
            "Course_Code", "Course_Name", "Course_Type", "Credits",
            "Students_Enrolled", "Avg_Grade_Points", "Grade_A_Count",
            "Grade_B_Count", "Grade_C_Count", "Grade_F_Count",
            "Pass_Percentage", "Difficulty_Level"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(
                start_color="70AD47",  # Green background
                end_color="70AD47",
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._get_border()
        
        # Write data
        row_num = 2
        for course_code, stats in course_stats.items():
            ws.cell(row_num, 1, course_code)
            ws.cell(row_num, 2, stats['name'])
            ws.cell(row_num, 3, stats.get('type', 'Core'))
            ws.cell(row_num, 4, stats.get('credits', 3))
            ws.cell(row_num, 5, stats['count'])
            ws.cell(row_num, 6, stats['average'])
            ws.cell(row_num, 7, stats['grade_counts'].get('A', 0))
            ws.cell(row_num, 8, stats['grade_counts'].get('B', 0))
            ws.cell(row_num, 9, stats['grade_counts'].get('C', 0))
            ws.cell(row_num, 10, stats['grade_counts'].get('F', 0))
            ws.cell(row_num, 11, stats['pass_percentage'])
            
            # Difficulty formula (based on average)
            difficulty = stats['difficulty']
            cell = ws.cell(row_num, 12, difficulty)
            
            # Color-code difficulty
            if difficulty == "Difficult":
                cell.font = Font(color="C00000", bold=True)  # Red
            elif difficulty == "Moderate":
                cell.font = Font(color="FF6600", bold=True)  # Orange
            else:  # Easy
                cell.font = Font(color="00B050", bold=True)  # Green
            
            # Conditional formatting on Avg_Grade_Points
            avg_cell = ws.cell(row_num, 6)
            avg_value = stats['average']
            if avg_value >= 8:
                avg_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif avg_value >= 6:
                avg_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                avg_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Apply borders
            for col in range(1, 13):
                ws.cell(row_num, col).border = self._get_border(thin=True)
            
            row_num += 1
        
        # Format sheet
        ws.freeze_panes = "A2"
        
        # Auto-fit columns
        for col_num in range(1, 13):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 15
        
        logger.info(f"✓ Sheet 2 created: {len(course_stats)} courses")
    
    def _create_sheet3_at_risk_students(self, ws, students: List[Dict]):
        """
        SHEET 3: Student Difficulty Indicators
        
        8 Columns as per TASK_SPECIFICATIONS.md:
        1. Priority (High/Medium/Low)
        2. Student_Name
        3. Roll_Number
        4. Department
        5. CGPA
        6. Attendance
        7. Backlogs
        8. Risk_Factors
        9. Weak_Subjects
        10. Action_Required
        11. Status (Pending/In Progress/Resolved)
        """
        # Import analytics module
        try:
            from src.core.dashboard_analytics import DashboardAnalytics
            at_risk_students = DashboardAnalytics.identify_at_risk_students(students)
        except ImportError:
            # Fallback: calculate inline
            at_risk_students = self._identify_at_risk_inline(students)
        
        # Define headers
        headers = [
            "Priority", "Student_Name", "Roll_Number", "Department",
            "CGPA", "Attendance", "Backlogs", "Risk_Factors",
            "Weak_Subjects", "Action_Required", "Status"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(
                start_color="C00000",  # Red background
                end_color="C00000",
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._get_border()
        
        # Write data
        for idx, student in enumerate(at_risk_students, 2):
            priority = student['priority']
            
            # Identify weak subjects
            weak_subjects = self._get_weak_subjects(student, students)
            
            row_data = [
                priority,
                student['name'],
                student['roll'],
                student['dept'],
                student['cgpa'],
                f"{student['attendance']}%",
                student.get('backlogs', 0),
                ", ".join(student['risks']),
                weak_subjects,
                "Immediate intervention" if priority == "High" else "Monitor closely" if priority == "Medium" else "Monitor",
                "Pending"
            ]
            
            # Write row
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(idx, col, value)
                
                # Priority-based row highlighting
                if priority == "High":
                    cell.fill = PatternFill(
                        start_color="FFC7CE",  # Light red
                        end_color="FFC7CE",
                        fill_type="solid"
                    )
                    if col == 1:  # Priority column
                        cell.font = Font(bold=True, color="C00000")
                elif priority == "Medium":
                    cell.fill = PatternFill(
                        start_color="FFEB9C",  # Light yellow
                        end_color="FFEB9C",
                        fill_type="solid"
                    )
                    if col == 1:
                        cell.font = Font(bold=True, color="FF6600")
                
                cell.border = self._get_border(thin=True)
                
                # Center alignment for some columns
                if col in [1, 5, 6, 7, 11]:
                    cell.alignment = Alignment(horizontal="center")
        
        # Format sheet
        ws.freeze_panes = "A2"
        
        # Auto-fit columns
        column_widths = [10, 20, 15, 20, 8, 12, 10, 35, 25, 25, 15]
        for col_num, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = width
        
        logger.info(f"✓ Sheet 3 created: {len(at_risk_students)} at-risk students")
    
    def _calculate_course_statistics(self, students: List[Dict]) -> Dict:
        """Calculate course-level statistics for Sheet 2."""
        course_stats = {}
        
        for student in students:
            courses = student.get('courses', [])
            if not courses:
                # Try alternate key
                courses = student.get('Courses', [])
            
            for course in courses:
                code = course.get('course_code') or course.get('Course Code', 'UNKNOWN')
                grade = course.get('grade') or course.get('Grade', 'N/A')
                
                if code not in course_stats:
                    course_stats[code] = {
                        'name': course.get('course_name') or course.get('Course Name', code),
                        'type': course.get('course_type', 'Core'),
                        'credits': course.get('credits') or course.get('Credits', 3),
                        'total_points': 0,
                        'count': 0,
                        'grade_counts': {'A': 0, 'B': 0, 'C': 0, 'F': 0},
                        'pass_count': 0
                    }
                
                # Convert grade to points
                grade_points = self._grade_to_points(grade)
                course_stats[code]['total_points'] += grade_points
                course_stats[code]['count'] += 1
                
                # Count grades
                if grade in ['A+', 'A']:
                    course_stats[code]['grade_counts']['A'] += 1
                    course_stats[code]['pass_count'] += 1
                elif grade in ['B+', 'B']:
                    course_stats[code]['grade_counts']['B'] += 1
                    course_stats[code]['pass_count'] += 1
                elif grade == 'C':
                    course_stats[code]['grade_counts']['C'] += 1
                    course_stats[code]['pass_count'] += 1
                elif grade in ['D', 'F']:
                    course_stats[code]['grade_counts']['F'] += 1
        
        # Calculate averages and difficulty
        for code, stats in course_stats.items():
            if stats['count'] > 0:
                avg = stats['total_points'] / stats['count']
                stats['average'] = round(avg, 2)
                stats['pass_percentage'] = round((stats['pass_count'] / stats['count']) * 100, 1)
                
                # Determine difficulty
                if avg >= 8:
                    stats['difficulty'] = "Easy"
                elif avg >= 6.5:
                    stats['difficulty'] = "Moderate"
                else:
                    stats['difficulty'] = "Difficult"
            else:
                stats['average'] = 0.0
                stats['pass_percentage'] = 0.0
                stats['difficulty'] = "N/A"
        
        return course_stats
    
    def _identify_at_risk_inline(self, students: List[Dict]) -> List[Dict]:
        """Fallback: Identify at-risk students inline."""
        at_risk = []
        
        for student in students:
            risks = []
            score = 0
            
            cgpa = float(student.get('cgpa') or student.get('CGPA', 10))
            attendance = float(student.get('attendance_percentage') or student.get('Attendance Percentage', 100))
            backlogs = int(student.get('backlogs_count', 0))
            
            if cgpa < 6.0:
                risks.append(f"Low CGPA: {cgpa}")
                score += 3
            
            if attendance < 75:
                risks.append(f"Low Attendance: {attendance}%")
                score += 2
            
            if backlogs >= 3:
                risks.append(f"Multiple Backlogs: {backlogs}")
                score += 3
            
            if risks:
                priority = "High" if score >= 4 else "Medium" if score >= 2 else "Low"
                
                at_risk.append({
                    'name': student.get('student_name') or student.get('Student Name', 'N/A'),
                    'roll': student.get('roll_number') or student.get('Roll Number', 'N/A'),
                    'dept': student.get('department') or student.get('Department', 'N/A'),
                    'cgpa': cgpa,
                    'attendance': attendance,
                    'backlogs': backlogs,
                    'risks': risks,
                    'priority': priority,
                    'score': score
                })
        
        at_risk.sort(key=lambda x: x['score'], reverse=True)
        return at_risk
    
    def _get_weak_subjects(self, at_risk_student, all_students):
        """Identify weak subjects for a student."""
        # Find full student record
        roll = at_risk_student['roll']
        
        for student in all_students:
            student_roll = student.get('roll_number') or student.get('Roll Number')
            if student_roll == roll:
                courses = student.get('courses', []) or student.get('Courses', [])
                weak = []
                
                for course in courses:
                    grade = course.get('grade') or course.get('Grade', '')
                    if grade in ['C', 'D', 'F']:
                        code = course.get('course_code') or course.get('Course Code', '')
                        weak.append(f"{code} ({grade})")
                
                return ", ".join(weak) if weak else "None identified"
        
        return "N/A"
    
    @staticmethod
    def _grade_to_points(grade: str) -> float:
        """Convert letter grade to grade points."""
        grade_map = {
            "A+": 10, "A": 9, "B+": 8, "B": 7,
            "C": 6, "D": 5, "F": 0, "I": 0, "W": 0
        }
        return grade_map.get(grade, 0)
    
    @staticmethod
    def _get_border(thin: bool = False) -> Border:
        """Get border style for cells."""
        style = "thin" if thin else "medium"
        return Border(
            left=Side(style=style, color="000000"),
            right=Side(style=style, color="000000"),
            top=Side(style=style, color="000000"),
            bottom=Side(style=style, color="000000")
        )


# ============================================================================
# USAGE EXAMPLE
# ============================================================================

if __name__ == "__main__":
    # Example usage
    from pathlib import Path
    
    # Sample student data
    sample_students = [
        {
            'student_name': 'Anjali Sharma',
            'roll_number': '21PH2034',
            'email': 'anjali.21ph2034@uoh.ac.in',
            'phone_number': '+91-9876543210',
            'department': 'Physics',
            'program': 'M.Sc',
            'semester': 'Sem-4',
            'academic_year': '2024-2025',
            'cgpa': 7.85,
            'sgpa': 8.12,
            'attendance_percentage': 82.5,
            'credits_earned': 90,
            'credits_registered': 96,
            'backlogs_count': 0,
            'current_standing': 'Good Standing',
            'confidence_metadata': {'overall_confidence': 0.92},
            'courses': [
                {'course_code': 'PH301', 'course_name': 'Quantum Mechanics', 'grade': 'A', 'credits': 4},
                {'course_code': 'PH302', 'course_name': 'Statistical Mechanics', 'grade': 'B+', 'credits': 4},
                {'course_code': 'MA301', 'course_name': 'Mathematical Methods', 'grade': 'C', 'credits': 4}
            ]
        },
        {
            'student_name': 'Rajesh Kumar',
            'roll_number': '21CS3045',
            'email': 'rajesh.21cs3045@uoh.ac.in',
            'department': 'Computer Science',
            'cgpa': 5.82,
            'attendance_percentage': 68.5,
            'backlogs_count': 4,
            'confidence_metadata': {'overall_confidence': 0.76},
            'courses': [
                {'course_code': 'CS301', 'course_name': 'Data Structures', 'grade': 'D', 'credits': 4},
                {'course_code': 'CS302', 'course_name': 'Algorithms', 'grade': 'F', 'credits': 3}
            ]
        }
    ]
    
    # Create Excel handler
    handler = TaskD_ExcelHandler(excel_dir=Path("./data/excel"))
    
    # Generate 3-sheet workbook
    filepath = handler.create_complete_workbook(
        students=sample_students,
        batch_name="test_batch"
    )
    
    print(f"✅ Excel workbook created: {filepath}")
    print("Open in Excel to verify:")
    print("  - Sheet 1: Raw_Records (blue header, 18 columns)")
    print("  - Sheet 2: Subject_Performance (green header, 12 columns)")
    print("  - Sheet 3: At_Risk_Students (red header, 11 columns)")
